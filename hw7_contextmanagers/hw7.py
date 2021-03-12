# Task 1
# У файлі task1.txt знаходиться текст субтитрів взятий з відео на ютубі. Текст складається з  міток часу і репліки яка
# була сказана в той момент часу.
# Причому репліка знаходиться в наступному рядку після мітки часу.
# Результатом виконнання завдання повинно бути:
# 1. словник елементами якого буде пара ключ:значення де ключ - мітка часу, значення - репліка в даний момент часу
# 2. файл в якому знаходиться текст з якого видалені всі мітки часу. всі субтитри повинні мати вигляд простого тексту.
# Це означає що окрім видалення міток часу, вам потрібно видалити переноси рядків
#
import pickle
import openpyxl


def subs_parse(filename):
    f_text = ''
    try:
        with open(filename, 'r') as file:  # 1
            f_list = file.read().split('\n')
        f_dict = {f_list[i]: f_list[i + 1] for i in range(0, len(f_list), 2)}
    except FileNotFoundError:
        print('File not found!!!')
        return None
    except IndexError:
        print('File is empty!!!!')
        return None

    for item in f_dict:  # 2
        f_text += f_dict[item]
    with open('output_' + filename, 'w') as output_file:
        output_file.write(f_text)

    return f_dict


print('Task1:\n', subs_parse('task1.txt'))

# Output:
# Task1:
# {'00:00': 'We all know you should take a sick day', '00:01': 'or a mental health day',
#  '00:03': 'or even a personal day.', '00:04': 'But a financial day is just as important.',
#  '00:07': '[Your Money and Your Mind with Wendy De La Rosa]',
#  '00:11': 'You deserve to take time to get your life in order', '00:14': "when you aren't on edge and under pressure.",
#  '00:16': "Employers, if you're watching this: you should also think about",
#  '00:20': 'giving your employees a financial health day.', '00:22': 'Throughout this series,',
#  '00:23': "I've shared ways that you can change your environment", '00:26': 'so you can spend less and save more.',
#  '00:28': 'But the secret ingredient for each tip has been time:',
#  '00:31': 'calling your credit card company to change your interest rate takes time;',
#  '00:35': 'deleting an app off your phone takes time.',
#  '00:37': "And if you haven't had time to put these changes into action,", '00:41': 'this is your chance.',
#  '00:42': 'Mark a day on your calendar right now', '00:45': 'to devote to reorganizing your finances.',
#  '00:47': 'Really commit to this day', '00:49': 'and make that day as productive as possible.',
#  '00:52': 'But here are the big headlines that are the most important to cover,',
#  '00:55': 'both from earlier episodes and beyond.', '00:58': 'So first I want you to focus on your fixed expenses.',
#  '01:01': 'Take the day to evaluate your bills.', '01:03': 'Can you truly afford your housing payment?',
#  '01:06': 'Is it time to start looking for a cheaper place?',
#  '01:08': 'Do you need to switch to a low-cost cell phone provider?',
#  '01:11': 'Do you need to trade in your car for something else?',
#  '01:14': 'This is the day where I really want you to think about', '01:16': 'those large fixed expenses',
#  '01:18': 'and make that one-time change.', '01:20': 'Number two: sign up for the boring-but-necessary stuff.',
#  '01:24': 'Do you have life insurance?', '01:25': 'No? Sign up now.',
#  '01:26': "Are you enrolled in your company's 401(k)?", '01:29': 'No? Sign up now.',
#  '01:30': 'Are you contributing enough to that 401(k)?', '01:32': 'No? Edit your contribution rate right now.',
#  '01:35': "Number three: if you haven't already,", '01:37': 'talk to your significant other about money.',
#  '01:39': 'Get on the same page; it matters.', '01:42': 'Visit the link below for a list of questions',
#  '01:44': 'to help get you started.', '01:46': 'Number four: create a singular savings goal.',
#  '01:49': 'You can start by setting up an automatic plan', '01:51': 'so that a portion of every paycheck',
#  '01:53': 'goes directly into your savings account.', '01:55': 'Number five: start paying off your debt every week,',
#  '01:58': 'not just every month.', '02:00': "Maybe it's your credit card debt",
#  '02:01': 'or your mortgage debt or whatever you have,',
#  '02:03': "you'll end up reducing more of your debt over the course of a year",
#  '02:07': 'if you pay that debt more often.', '02:09': 'Number six: renegotiate your credit card interest rate',
#  '02:12': 'and change your payment due date.', '02:14': 'Get on the phone, call your credit card company',
#  '02:17': 'and make your credit card company work for you.', '02:19': 'Number seven: use technology to your advantage.',
#  '02:22': 'Redesign your online environment', '02:24': 'by unsubscribing from all your shopping newsletters',
#  '02:28': 'and install an ad blocker.', '02:29': "You can't spend on what you can't see.",
#  '02:32': 'Number eight: delete those distracting delivery apps.', '02:35': 'Number nine:',
#  '02:37': 'now, my tips are not all about cutting back spending.',
#  '02:39': 'I want you to spend on things that increase your happiness.',
#  '02:43': 'So focus on experiences, on spending time with others',
#  '02:46': 'and on expenditures that help save you time,', '02:49': 'like hiring someone to clean your house',
#  '02:51': 'or mow your lawn.', '02:53': 'That will save you time and boost your happiness.',
#  '02:56': 'Number 10: last but certainly not least,', '02:59': 'schedule another financial health day',
#  '03:01': 'for a few weeks later.', '03:02': "Chances are, you'll need to revisit some of the things",
#  '03:05': 'you started today.', '03:06': 'I know all of this may seem tedious or boring,',
#  '03:08': "but it's better to make these vital changes", '03:11': 'when you have the time to think about them,',
#  '03:13': 'and not under the stressful conditions', '03:15': 'when we usually make these very important decisions.',
#  '03:18': 'Not every day can be a spa day,', '03:20': 'but I think a financial day',
#  '03:22': 'will leave you feeling soothed and pampered.'}


# Task 2
# в файлі task2 збережений список, відкрийте цей файл, прочитайте вміст, і знайдіть середнє арифметичне чисел
# що знаходяться в списку
#

def average(filename):
    try:
        with open(filename, 'rb') as file:
            f_list = pickle.load(file)
        sum_aver = sum(f_list) / len(f_list)
    except FileNotFoundError:
        print('File not found!!!')
        return None
    except EOFError:
        print('File is corrupted!!!')
        return None
    return sum_aver


print('Task2:\n', average('task2'))

# Output:
# Task2:
#  187.52631578947367


# Task 3
# Використовуючи openpyxl (або будь-яку іншу зручну для вас бібліотеку),
# напишіть контекстний менеджер для роботи з ексель.
# Даний менеджер повинен бути аналогом методу open()


class Excel:
    def __init__(self, filename):
        self.__name = filename
        try:
            self.file_obj = openpyxl.load_workbook(filename=filename)
        except FileNotFoundError:  # create new file if file not exist
            self.__temp = openpyxl.Workbook()
            self.__temp.save(self.__name)
            self.file_obj = openpyxl.load_workbook(filename=filename)

    def __enter__(self):
        return self.file_obj

    def __exit__(self, exc_type, exc_value, exc_traceback):
        self.file_obj.save(self.__name)
        self.file_obj.close()


with Excel('task3_file.xlsx') as wb:
    ws = wb.active
    ws['A1'] = 123
    ws['A2'] = 'asd'
    print('Task 3:')
    print(ws['A1'].value)
    print(ws['A2'].value)


# Output:
# Task 3:
# 123
# asd